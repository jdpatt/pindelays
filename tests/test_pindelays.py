from pathlib import Path

import pytest
from openpyxl import Workbook

from pindelays.pindelays import (
    generate_cadence,
    generate_mentor,
    get_column,
    parse_excel_file,
)


@pytest.fixture(scope="session")
def workbook():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Pin Name"
    ws["B1"] = "Delay"
    ws["A2"] = "A1"
    ws["B2"] = "10"
    ws["A3"] = "A2"
    ws["B3"] = "12"
    return wb


@pytest.mark.usefixtures("workbook")
class TestPinDelays(object):
    def test_get_column(self, workbook):
        """ If the cell exists return the position otherwise return None """
        worksheet = workbook.active
        assert get_column("Delay", worksheet) == 2
        assert get_column("foo", worksheet) == 0

    def test_parse_excel_file(self, workbook):
        assert parse_excel_file(workbook) == {"A1": "10", "A2": "12"}
        assert parse_excel_file(workbook, 1, 2)

    def test_parse_excel_file_value_error(self, workbook):
        workbook.active["B2"] = ""  # Missing delay for pin A1
        with pytest.raises(ValueError):
            parse_excel_file(workbook)

    def test_generate_cadence(self, workbook):
        delays = {"A1": "10", "A2": "12"}
        generate_cadence("U1", "test_package", "MIL", delays)
        with open(Path().cwd().joinpath("test_package.csv")) as cadence:
            assert cadence.read() == (
                "PIN DELAY\n"
                "REFDES\tU1\n"
                "DEVICE\ttest_package\n\n"
                "A1\t10\tMIL\n"
                "A2\t12\tMIL\n"
            )

    def test_generate_mentor_mils(self, workbook):
        delays = {"A1": "10", "A2": "12"}
        generate_mentor("test_partnumber", "mil", delays)
        with open(Path().cwd().joinpath("PinPkgLengths.txt")) as mentor:
            assert mentor.read() == (
                "UNITS th\n" "PART_NUMBER test_partnumber\n" "A1 10\n" "A2 12\n"
            )

    def test_generate_mentor_delay(self, workbook):
        delays = {"A1": "10", "A2": "12"}
        generate_mentor("test_partnumber", "ps", delays)
        with open(Path().cwd().joinpath("PinPkgDelays.txt")) as mentor:
            assert mentor.read() == (
                "UNITS ps\n" "PART_NUMBER test_partnumber\n" "A1 10\n" "A2 12\n"
            )
