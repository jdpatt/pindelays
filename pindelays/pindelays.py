"""For today's high speed designs, one must take into account the internal length or delay of a pin.
This python program takes an excel file and produces a pin delay file that is correctly formatted
for your EDA tool set. """
from pathlib import Path
from typing import Any, Dict

import click
from openpyxl import Workbook, load_workbook


def get_column(name: str, worksheet) -> int:
    """Search a row of cells for the string.

    Args:
        name: The text to search for
        columns: The list or generator of columns in the excel sheet

    Returns:
        Either returns the column number or returns 0 if no column matched the name

    """
    for rows in worksheet.iter_rows(min_row=1, max_row=1, min_col=1):
        for column in rows:
            if column.value == name:
                return column.col_idx
    return 0


def parse_excel_file(workbook: Workbook, pin_col=None, delay_col=None) -> Dict[str, Any]:
    """Read in excel and get the pin number and internal length

    The excel file must have a header row with the cells "Pin Name" and "Delay".  It does not
    matter which column they are in.

    Args:
        excel_file: The excel file to open and read from

    """
    sheet = workbook.active
    delay_dict = dict()
    try:
        pin_col = pin_col or get_column("Pin Name", sheet)
        delay_col = delay_col or get_column("Delay", sheet)
        for excel_row in range(2, sheet.max_row + 1):
            pin = str(sheet.cell(row=excel_row, column=pin_col).value)
            delay = str(sheet.cell(row=excel_row, column=delay_col).value)
            if not all([pin, delay]):
                raise ValueError
            delay_dict.update({pin: delay})
    except (ValueError, KeyError, UnboundLocalError) as error:
        print(error)
        raise
    return delay_dict


def generate_mentor(partnumber: str, unit: str, delays: Dict) -> None:
    """This function generates a text file that can be imported in the Constraint Manager tool.

    Example:
        UNITS <value> th
        PART_NUMBER <part_number>
        <pin_number> <value>

    Args:
        partnumber: The part number to apply these delays
        delays: The data read in from the excel file

    """
    if unit == "mil":
        filename = "PinPkgLengths.txt"
        unit = "th"
    else:
        filename = "PinPkgDelays.txt"
    with open(filename, "w") as output:
        output.write(f"UNITS {unit}\n")
        output.write(f"PART_NUMBER {partnumber}\n")
        for key, value in delays.items():
            output.write(f"{key} {value}\n")


def generate_cadence(ref: str, package: str, unit: str, delays: Dict) -> None:
    """This function generates a text file that can be imported into Allergo if you are using the
    high speed license. Allergo applies delays individual vs against all part numbers that match
    like mentor.  UNITS MIL can be a header row that applies to everything or you can list unit for
    every row.  This does the later.

    Example:
        [PIN DELAY]
        [RefDes    <refdes>]
        [DEVICE    <package name>]
        [UNITS     <mks units>]
        <Pin number>    <delay value> <...>

    Args:
        ref:  The reference designator to apply the delays
        package: The cadence source package
        unit: The unit of the delays in either MIL or NS
        delays: The data read in from the excel file

    """
    with open(f"{package}.csv", "w") as output:
        output.write("PIN DELAY\n")
        output.write(f"REFDES\t{ref}\n")
        output.write(f"DEVICE\t{package}\n")
        output.write("\n")
        for key, value in delays.items():
            output.write(f"{key}\t{value}\t{unit.upper()}\n")


@click.command(context_settings=dict(help_option_names=["-h", "--help"]))
@click.argument("excel_file", nargs=-1)
@click.argument("output_type", type=click.Choice(["cadence", "mentor"]), default="cadence")
@click.option(
    "--partnumber",
    "-p",
    type=str,
    default="dummy_part",
    help="Part number [Only used in mentor]",
)
@click.option(
    "--package",
    "-d",
    type=str,
    default="dummy_package",
    help="Device Package [Only used in cadence]",
)
@click.option("--refdes", "-r", default="U1", type=str, help="RefDes [Only used in cadence]")
@click.option(
    "--units",
    "-u",
    type=click.Choice(["ns", "ps", "mil"]),
    default="ns",
    help="Units",
)
@click.version_option()
def pindelay(excel_file, output_type, partnumber, package, refdes, units):
    """For today's high speed designs, one must take into account the internal length or delay of a pin.
    This python program takes an excel file and produces a pin delay file that is correctly formatted
    for your EDA tool set."""
    for file_to_parse in excel_file:
        print(f"Reading in Excel File: {Path(file_to_parse)}")
        part_delays = parse_excel_file(load_workbook(file_to_parse, data_only=True))
        if output_type == "cadence":
            generate_cadence(refdes, package, units, part_delays)
            print("Cadence File Generated")
        if output_type == "mentor":
            generate_mentor(partnumber, units, part_delays)
            print("Mentor File Generated")


if __name__ == "__main__":
    pindelay()
